import pandas as pd
import spacy
from spacy.matcher import PhraseMatcher
from openpyxl.styles import PatternFill

# Load English language model
nlp = spacy.load("en_core_web_lg")

# Load the Excel file
df = pd.read_excel('incidents.xlsx')

# Define severity patterns
severity_patterns = {
    'L3': ['outage', 'failure', 'breach', 'ransomware', 'unauthorized access', 
           'data loss', 'downtime', 'critical error', 'system crash'],
    'L2': ['degradation', 'timeout', 'high latency', 'performance issue', 
           'authentication failure', 'service disruption'],
    'L1': ['warning', 'notification', 'minor issue', 'configuration alert']
}

# Initialize PhraseMatcher
matcher = PhraseMatcher(nlp.vocab)
for severity, phrases in severity_patterns.items():
    patterns = [nlp(text) for text in phrases]
    matcher.add(severity, patterns)
# Define classification rules
def classify_severity(row):
    description = str(row['Description']).lower()
    root_cause = str(row['RootCause']).lower()
    resolution_time = row['ResolutionTime(min)']
    impact = str(row['ImpactLevel']).lower()
    
    # L3 Criteria (Complex)
    if (resolution_time > 180 or 
        'complex' in impact or
        any(x in description for x in ['outage', 'failure', 'halted', 'breach', 'ransomware', 'unauthorized']) or
        any(x in root_cause for x in ['controller failure', 'regional failure', 'privilege escalation'])):
        return 'L3'
    
    # L2 Criteria (Major)
    elif (60 <= resolution_time <= 180 or
          'high' in impact or
          any(x in description for x in ['timeout', 'degraded', 'slow', 'error', 'attack']) or
          any(x in root_cause for x in ['memory leak', 'ssl', 'misconfiguration'])):
        return 'L2'
    
    # L1 Criteria (Minor)
    else:
        return 'L1'
# Enhanced classification function
def classify_with_nlp(row):
    doc = nlp(str(row['Description']).lower() + " " + str(row['RootCause']).lower())
    matches = matcher(doc)
    
    # Default to rules-based classification
    severity = classify_severity(row)  # From previous rules-based function
    
    # Override with NLP findings if high confidence
    if matches:
        _, start, end = matches[0]  # Take first match
        span = doc[start:end]
        if span.text in severity_patterns['L3']:
            severity = 'L3'
        elif span.text in severity_patterns['L2'] and severity != 'L1':
            severity = 'L2'
    
    # Add NLP insights
    entities = [(ent.text, ent.label_) for ent in doc.ents]
    return pd.Series([severity, str(entities)])

# Apply classification
df[['Severity', 'NLP_Insights']] = df.apply(classify_with_nlp, axis=1)

# Entity recognition statistics
tech_entities = ['ORG', 'PRODUCT', 'TECH']  # spaCy entity labels
df['Tech_Components'] = df['Description'].apply(
    lambda x: [ent.text for ent in nlp(x).ents if ent.label_ in tech_entities]
)
#df = df.map(lambda x: str(x).replace('[', '').replace(']', ''))
df = df.map(lambda x: str(x).replace('[]', 'None'))

with pd.ExcelWriter('tickets.xlsx') as writer:
    # for col in df.select_dtypes(include='object').columns:
    #     df[col] = df[col].str.replace(r'[\[\]]', '', regex=True)
    
    df.to_excel(writer, index=False)
    
    # Add color coding
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # Conditional formatting
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C")
    
    for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=7, max_col=7):
        for cell in row:
            if cell.value == 'L3':
                cell.fill = red_fill
            elif cell.value == 'L2':
                cell.fill = yellow_fill